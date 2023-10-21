import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
def stock_name(StockName):
    # 設置 index constant，數字代表我們要的資料在 list 的位置
    TARGET_TABLE_INDEX = 1
    STOCK_NO_INDEX = 2
    STOCK_NAME_INDEX = 3
    STOCK_INDUSTRY_INDEX = 6
    # JSON settings
    TITLE = "stock"
    JSON_INDENT = 4

    # 送出 HTTP Request
    url = "https://isin.twse.com.tw/isin/class_main.jsp"
    res = requests.get(url, params={
        "market": "1",
        "issuetype": "1",
        "Page": "1",
        "chklike": "Y"
    })

    # 處理編碼，使用預設 utf-8 的話 res.text 的內容會有亂碼
    res.encoding = "Big5"
    res_html = res.text

    # Parse
    soup = BeautifulSoup(res_html, "lxml")

    # 因為這個 HTML 裡面有兩張 table
    # 所以我們 find_all("table") 回傳的 list 的 length 會是 2
    # 而我們要的資料在第二張
    tr_list = soup.find_all("table")[TARGET_TABLE_INDEX].find_all("tr")

    # tr_list 的第一個是 item 是欄位名稱
    # 我們這邊用不到所以 pop 掉
    tr_list.pop(0)

    # 開始處理資料
    result = []
    for tr in tr_list:

        td_list = tr.find_all("td")

        # 股票代碼
        stock_no_val = td_list[STOCK_NO_INDEX].text

        # 股票名稱
        stock_name_val = td_list[STOCK_NAME_INDEX].text

        # 股票產業類別
        stock_industry_val = td_list[STOCK_INDUSTRY_INDEX].text

        # 整理成 dict 存起來
        result.append({
            stock_name_val: stock_no_val,
            # "stockNo": stock_no_val,
            # "stockName": stock_name_val,
            #"stockIndustry": stock_industry_val
        })

    for dic in result:
        if StockName in dic:
            return f"{dic[StockName]}"
        
    return f"{StockName}"

def extract(assets):
    counter = -1
    dict1 = {}
    list1 = []
    key_list = []
    for i in assets:
        content_list = []
        for j in i:
            if j.text != "\n":
                content_list.append(j.text)
            if counter %  9  == 0:
                # print(j.text,end=" ")
                key_list.append(j.text)
            counter+=1
        content_list.pop(0)
        list1.append(content_list)

    for i in range(len(key_list)):
        dict1[key_list[i]] = list1[i]
        
    return dict1
    # for key,value in dict1.items():
    #     print(key,end=" ")
    #     for i in value:
    #         print(i,end=" ")

a = input()
url = "https://histock.tw/stock/"+stock_name(a)+"/%E8%82%A1%E6%9D%B1%E6%AC%8A%E7%9B%8A"

response = requests.get(url)
soup = BeautifulSoup(response.content, "lxml")
total_assets = soup.find_all("tr",class_ = "alt-row")
year = soup.find("th",class_="w1 f13 w80")
title = soup.find_all("th",class_="w1 f13 w95")
q1_assets = soup.find_all("tr", align="center")
wb = openpyxl.Workbook()

ws = wb.active
ws.title = a


n = 2
j = 2
ws.cell(1,1).value = "年度 季"
for i in title:
    ws.cell(1,j).value = i.text
    j+=1
for key1, value1 in extract(q1_assets).items():
    ws.cell(n, 1).value = key1
    for i, v in enumerate(value1, start=2):
        ws.cell(n, i).value = v
    n += 1

list2023  =[] 
currentLiabiltiy2023 = 0 
currentLiabiltiy2022 = 0
currentLiabiltiy2021 = 0
currentLiabiltiy2020 = 0
longLiability2023 = 0
longLiability2022 = 0
longLiability2021 = 0
longLiability2020 = 0
other2023 = 0
other2022 = 0
other2021 = 0
other2020 = 0
totalLiability2023 = 0
totalLiability2022 = 0
totalLiability2021 = 0
totalLiability2020 = 0
netValue2023 = 0
netValue2022 = 0
netValue2021 = 0
netValue2020 = 0
totalAsset2023 = 0
totalAsset2022 = 0
totalAsset2021 = 0
totalAsset2020 = 0


data = extract(q1_assets)
for key1,value1 in data.items():
    if("2023" in key1):
        currentLiabiltiy2023+=int(value1[0].replace(",",""))
        longLiability2023+=int(value1[1].replace(",",""))
        other2023+=int(value1[2].replace(",",""))
        totalLiability2023+=int(value1[3].replace(",",""))
        netValue2023+=int(value1[4].replace(",",""))
        totalAsset2023+=int(value1[5].replace(",",""))
    elif("2022" in key1):
        currentLiabiltiy2022+=int(value1[0].replace(",",""))
        longLiability2022+=int(value1[1].replace(",",""))
        other2022+=int(value1[2].replace(",",""))
        totalLiability2022+=int(value1[3].replace(",",""))
        netValue2022+=int(value1[4].replace(",",""))
        totalAsset2022+=int(value1[5].replace(",",""))
    elif("2021" in key1):
        currentLiabiltiy2021+=(int(value1[0].replace(",","")))
        longLiability2021+=int(value1[1].replace(",",""))
        other2021+=int(value1[2].replace(",",""))
        totalLiability2021+=int(value1[3].replace(",",""))
        netValue2021+=int(value1[4].replace(",",""))
        totalAsset2021+=int(value1[5].replace(",",""))
    elif("2020" in key1):
        currentLiabiltiy2020+=(int(value1[0].replace(",","")))
        longLiability2020+=int(value1[1].replace(",",""))
        other2020+=int(value1[2].replace(",",""))
        totalLiability2020+=int(value1[3].replace(",",""))
        netValue2020+=int(value1[4].replace(",",""))
        totalAsset2020+=int(value1[5].replace(",",""))

yearList = [[currentLiabiltiy2023,longLiability2023,other2023,
             totalLiability2023,netValue2023,totalAsset2023],
             [currentLiabiltiy2022,longLiability2022,other2022,
             totalLiability2022,netValue2022,totalAsset2022],
             [currentLiabiltiy2021,longLiability2021,other2021,
             totalLiability2021,netValue2021,totalAsset2021],
             [currentLiabiltiy2020,longLiability2020,other2020,
             totalLiability2020,netValue2020,totalAsset2020]]
j = 10
ws.cell(1,9).value = "年度"
ws.cell(2,9).value = "2023"
ws.cell(3,9).value = "2022"
ws.cell(4,9).value = "2021"
ws.cell(5,9).value = "2020"

for i in title:
    ws.cell(1,j).value = i.text
    j+=1
a = 2
for i in yearList:
    b = 10
    for j in i:
        ws.cell(a,b).value = j
        b+=1
    a+=1

wb.save(filename='data.xlsx')